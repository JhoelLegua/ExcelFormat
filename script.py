import pandas as pd
import glob
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime



# Función para formatear fechas
def formatear_fecha(fecha):
    if pd.isna(fecha) or fecha == '' or fecha is None:
        return ''
    
    try:
        # Si ya es un objeto datetime o Timestamp
        if isinstance(fecha, (pd.Timestamp, datetime)):
            return fecha.strftime('%-d/%-m/%Y') if os.name != 'nt' else fecha.strftime('%d/%m/%Y').lstrip('0').replace('/0', '/')
        
        fecha_str = str(fecha).strip()
        
        # Si ya está en formato D/M/YYYY o DD/MM/YYYY, verificar y ajustar
        if '/' in fecha_str and len(fecha_str.split('/')) == 3:
            partes = fecha_str.split('/')
            try:
                dia = int(partes[0])
                mes = int(partes[1])
                año = int(partes[2])
                if 1 <= dia <= 31 and 1 <= mes <= 12 and año > 1900:
                    return f"{dia}/{mes}/{año}"
            except ValueError:
                pass
        
        # Intentar parsear diferentes formatos de fecha
        formatos = ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']
        
        for formato in formatos:
            try:
                fecha_obj = datetime.strptime(fecha_str, formato)
                return f"{fecha_obj.day}/{fecha_obj.month}/{fecha_obj.year}"
            except ValueError:
                continue
        
        # Si es un timestamp de Excel (número)
        if fecha_str.replace('.', '').replace('-', '').isdigit():
            fecha_excel = float(fecha_str)
            # Excel cuenta días desde 1900-01-01 (con ajuste por bug del año 1900)
            fecha_obj = datetime(1899, 12, 30) + pd.Timedelta(days=fecha_excel)
            return f"{fecha_obj.day}/{fecha_obj.month}/{fecha_obj.year}"
        
        # Si no se puede convertir, devolver el valor original
        return fecha_str
        
    except Exception as e:
        print(f"   Error formateando fecha '{fecha}': {e}")
        return str(fecha) if fecha else ''

# ---------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------
# Directorio base: junto al ejecutable cuando está congelado, o junto al script en desarrollo
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

INPUT_DIR = os.path.join(BASE_DIR, 'input')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Incluir .xls y .xlsx
ruta_archivos = os.path.join(INPUT_DIR, '*.xls*')
archivo_salida = os.path.join(OUTPUT_DIR, 'Planilla_Unificada_Final.xlsx')

# ---------------------------------------------------------
# LÓGICA
# ---------------------------------------------------------
print("Iniciando unificacion estricta...")

archivos = glob.glob(ruta_archivos)
archivos.sort() # Ordenamos para que se unan en orden (1, 2, 3...)

lista_dataframes = []

for i, archivo in enumerate(archivos):
    nombre = os.path.basename(archivo)
    print(f"Procesando: {nombre}")
    
    try:
        # PASO CLAVE: 'header=2'
        # Esto le dice a Python: "Salta las filas 0 y 1. La fila 2 (la tercera visualmente) son los títulos".
        # dtype=object: Importante para leer "UE: 1" sin que se rompa por esperar números.
        df = pd.read_excel(archivo, header=2, dtype=object)
        
        # Limpieza de columnas (quitar espacios en blanco en los nombres de la fila amarilla)
        df.columns = df.columns.str.strip().str.replace('\n', ' ')
        
        # Estrategia simple: revisar si hay una columna que podría ser "Item"
        # Buscar en las primeras columnas una que tenga números secuenciales
        posible_item_col = None
        for i, col in enumerate(df.columns[:5]):  # Solo revisar las primeras 5 columnas
            if 'Unnamed' in str(col):
                sample_data = df[col].dropna().head(10)
                if len(sample_data) > 0:
                    # Verificar si son números que podrían ser items (1, 2, 3, etc.)
                    try:
                        numeric_vals = pd.to_numeric(sample_data, errors='coerce').dropna()
                        if len(numeric_vals) >= 3 and numeric_vals.min() >= 1:
                            posible_item_col = col
                            break
                    except:
                        continue
        
        # Eliminar columnas "Unnamed" EXCEPTO la que identificamos como posible Item
        columnas_a_mantener = []
        for col in df.columns:
            if 'Unnamed' in str(col):
                if col == posible_item_col:
                    columnas_a_mantener.append(col)
            else:
                columnas_a_mantener.append(col)
        
        df = df[columnas_a_mantener]
        
        # Renombrar la columna Item si la encontramos
        if posible_item_col and posible_item_col in df.columns:
            df = df.rename(columns={posible_item_col: 'Item'})
            print(f"   Columna Item detectada y renombrada: {posible_item_col} -> Item")
        
        # Eliminar columnas completamente vacías
        df = df.dropna(axis=1, how='all')
        
        # Agregar columna Item si no existe (después de Fecha de Nac.)
        if 'Item' not in df.columns:
            # Buscar la posición de "Fecha de Nac." o columnas similares
            fecha_col_pos = None
            for i, col in enumerate(df.columns):
                if 'fecha' in str(col).lower() and 'nac' in str(col).lower():
                    fecha_col_pos = i
                    break
            
            if fecha_col_pos is not None:
                # Insertar columna Item después de Fecha de Nac.
                columnas_ordenadas = list(df.columns)
                columnas_ordenadas.insert(fecha_col_pos + 1, 'Item')
                
                # Crear el nuevo dataframe con la columna Item vacía
                df['Item'] = ''  # Columna vacía por defecto
                df = df[columnas_ordenadas]
                print(f"   Columna 'Item' agregada después de la columna de fecha")
            else:
                # Si no encontramos fecha, agregar al final
                df['Item'] = ''
                print(f"   Columna 'Item' agregada al final")
        
        # Formatear fechas de nacimiento al formato DD/MM/YYYY
        for col in df.columns:
            if 'nacimiento' in str(col).lower() or 'fecha' in str(col).lower():
                print(f"   Procesando fechas en columna: {col}")
                df[col] = df[col].apply(lambda x: formatear_fecha(x))
        
        # Debug: Verificar si existe "Totales por Estructura Programática:"
        tiene_totales = df.astype(str).apply(lambda x: x.str.contains('Totales por Estructura Programática:', na=False)).any().any()
        if tiene_totales:
            print(f"   Encontrado 'Totales por Estructura Programatica:' en {nombre}")
        
        # Eliminar SOLO filas que contengan "Total Funcionarios :" Y que NO contengan "Totales por Estructura Programática:"
        mask_total_funcionarios = df.astype(str).apply(lambda x: x.str.contains('Total Funcionarios :', na=False)).any(axis=1)
        mask_totales_estructura = df.astype(str).apply(lambda x: x.str.contains('Totales por Estructura Programática:', na=False)).any(axis=1)
        
        # Mantener filas que NO sean "Total Funcionarios :" O que SÍ sean "Totales por Estructura Programática:"
        df = df[~mask_total_funcionarios | mask_totales_estructura]
        
        # Eliminar filas completamente vacías DESPUÉS de los filtros de texto
        df = df.dropna(how='all')
        
        # Agregamos este bloque a la lista
        lista_dataframes.append(df)
        
    except Exception as e:
        print(f"Error leyendo {nombre}: {e}")

# ---------------------------------------------------------
# GENERACIÓN DEL ARCHIVO FINAL
# ---------------------------------------------------------
if lista_dataframes:
    print("Uniendo bloques...")
    
    # pd.concat une todos los dataframes verticalmente.
    # Al usar los mismos nombres de columnas (de la fila 3), se alinean perfectamente.
    df_final = pd.concat(lista_dataframes, ignore_index=True)
    
    # Exportamos a Excel
    # index=False para que no agregue una columna de números a la izquierda
    df_final.to_excel(archivo_salida, index=False)
    
    # Formatear la primera fila (headers)
    print("Aplicando formato a headers...")
    wb = load_workbook(archivo_salida)
    ws = wb.active
    
    # Definir estilos
    fill_verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_naranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    font_bold = Font(bold=True)
    
    # Aplicar formato a la primera fila
    for cell in ws[1]:
        if cell.value:
            cell.value = str(cell.value).upper()  # Convertir a mayúsculas
            cell.fill = fill_verde  # Fondo verde
            cell.font = font_bold   # Texto en negrita
    
    # Colorear filas que contengan los campos específicos
    campos_especiales = ["UE:", "Programa:", "Proyecto:", "Actividad:", "Fuente:", "Organismo:", "Totales por Estructura Programática:"]
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Empezar desde fila 2 (después del header)
        # Obtener valores de la fila para análisis
        valores_fila = [cell.value for cell in row]
        
        # Detectar filas de totales: tienen números pero las primeras columnas están vacías o con pocos datos
        es_fila_totales = False
        if len(valores_fila) > 5:  # Asegurar que hay suficientes columnas
            # Verificar si las primeras 3-4 columnas están mayormente vacías pero hay números después
            primeras_vacias = sum(1 for val in valores_fila[:4] if val is None or str(val).strip() == '')
            tiene_numeros = any(isinstance(val, (int, float)) and val != 0 for val in valores_fila[4:])
            
            # Si las primeras columnas están vacías y hay números, probablemente es una fila de totales
            if primeras_vacias >= 3 and tiene_numeros:
                es_fila_totales = True
        
        # Aplicar colores
        if es_fila_totales:
            # Colorear fila de totales de NARANJA
            for cell_in_row in row:
                cell_in_row.fill = fill_naranja
        else:
            # Verificar campos especiales para AMARILLO
            for cell in row:
                if cell.value and any(campo in str(cell.value) for campo in campos_especiales):
                    for cell_in_row in row:
                        cell_in_row.fill = fill_amarillo
                    break
    
    wb.save(archivo_salida)
    
    print(f"\nListo! Archivo generado: {archivo_salida}")
    print(f"   Se eliminaron las 2 primeras filas de cada archivo.")
    print(f"   Se conserva 1 sola cabecera (Fila Amarilla).")
    print(f"   Headers formateados: VERDE, NEGRITA, MAYÚSCULAS.")
    print(f"   Filas con campos especiales coloreadas de AMARILLO.")
    print(f"   Filas de totales generales coloreadas de NARANJA.")
    print(f"   Filas totales unificadas: {len(df_final)}")
    
else:
    print("No se encontraron archivos para procesar.")